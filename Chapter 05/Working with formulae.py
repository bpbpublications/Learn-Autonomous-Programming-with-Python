from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.formula.translate import Translator
objWorkbook = load_workbook("Student Table.xlsx")
objWorksheet = objWorkbook.worksheets[0]
objWorksheet['B5']="=ROUND(AVERAGE(B2:B4),0)"
row=5
col=3
translator = Translator("=ROUND(AVERAGE(B2:B4),0)","B5")

for i in range(4):
  str_col_letter = objWorksheet.cell(row,col).column_letter
  objWorksheet[str_col_letter + '5'] =translator.translate_formula(str_col_letter + "5")
  objWorksheet.cell(row,col).alignment = Alignment(horizontal='center',vertical='center')
  col+=1
objWorkbook.save("Student Table.xlsx")
