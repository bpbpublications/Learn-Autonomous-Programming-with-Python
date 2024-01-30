from openpyxl import load_workbook  
objWorkbook = load_workbook("Learn_Excel_Automation.xlsx")  
objWorksheet = objWorkbook.worksheets[0]

objWorksheet.cell(3,1,'This is my second task in openpyxl')
objWorkbook.save("Learn_Excel_Automation.xlsx")



