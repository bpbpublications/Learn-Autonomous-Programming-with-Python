from openpyxl import load_workbook 
from openpyxl.styles import Alignment   
objWorkbook = load_workbook("Learn_Excel_Automation.xlsx")  
objWorksheet = objWorkbook.worksheets[0]

objWorksheet.merge_cells('A1:D2')  
cell = objWorksheet.cell(1,1)  
cell.alignment = Alignment(horizontal='center', vertical='center') 
objWorkbook.save("Learn_Excel_Automation.xlsx")
