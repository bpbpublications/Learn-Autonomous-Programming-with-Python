#xlsxwriter

import xlsxwriter as xw
objWorkbook = xw.Workbook('Stock Prices.xlsx')
objWorksheet = objWorkbook.add_worksheet('Prices')

dict_Stock_Prices = {'A':[357,	457,	187,	831,	779,	338,	129,	508,	407,	748,	511,	609],\
                     'B':[14908,	13408,	17103,	18886,	19828,	12098,	17080,	16850,	15023,	12405,	15469,	13800],\
                     'C':[60,64,54,	93,	87,	74,	96,	92,	83,85,70,88],\
                     'D':[1667,	1962,	1845,	1535,	1753,	1767,	1551,	1893,	1715,	1707,	1627,	1532],\
                     'E':[2181,	2333,	2265,	2274,	2739,	2601,	2569,	2520,	2744,	2234,	2836,	2230]}

month_list = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
stock_list = ['A','B','C','D','E']

objFormat_1=objWorkbook.add_format({'border':1})
objFormat_2=objWorkbook.add_format({'border':2})

objWorksheet.write_column('A2', stock_list,objFormat_2)
objWorksheet.write_row('B1', month_list, objFormat_2)

row=1
for stock in dict_Stock_Prices:
    stock_prices = dict_Stock_Prices[stock]
    objWorksheet.write_row(row,1, stock_prices, objFormat_1)
    row+=1
    
objWorkbook.close()   


#xlwings
import xlwings as xlw
obj_excel_app = xlw.App(visible=False)
objWorkbook = xlw.Book('Stock Prices.xlsx') 
objWorksheet = objWorkbook.sheets[0]
formula = '=ROUND(AVERAGE(B2:M2),0)'
objWorksheet.range("N2,N3,N4,N5,N6").formula = formula
objWorkbook.save()
objWorkbook.close()
obj_excel_app.kill()


#openpyxl
import openpyxl
from openpyxl import load_workbook 
from openpyxl.chart import PieChart,Reference
from openpyxl.chart.layout import Layout, ManualLayout
objWorkbook = openpyxl.load_workbook('Stock Prices.xlsx')
objWorksheet = objWorkbook.worksheets[0]
labels = Reference(objWorksheet, min_col = 1, min_row = 2, max_row = 6)
data = Reference(objWorksheet, min_col = 14, min_row = 2, max_row = 6)
chart = PieChart()
chart.add_data(data, titles_from_data = False)
chart.set_categories(labels)
chart.title = " PIE-CHART "
objWorksheet.add_chart(chart, "C10")
chart = objWorksheet._charts[0]
chart.title = "Average Price of Stocks" 
objWorkbook.save("Stock Prices.xlsx")

