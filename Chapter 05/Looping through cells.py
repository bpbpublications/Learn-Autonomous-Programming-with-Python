from openpyxl import load_workbook 
from openpyxl.styles import Alignment   

objWorkbook = load_workbook("Student Table.xlsx")  
objWorksheet = objWorkbook.worksheets[0]

dict_scores = {'A':{'Mathematics':45,'Science':55,'English':75},\
               'B':{'Mathematics':75,'Science':87,'English':87},\
               'C':{'Mathematics':90,'Science':95,'English':58},\
               'D':{'Mathematics':100,'Science':28,'English':80},\
               'E':{'Mathematics':35,'Science':75,'English':90}}
col=2    
for key in dict_scores:
    student = dict_scores[key]
    row=2
    for subject in student:
        score = student[subject]
        objWorksheet.cell(row,col).value = score
        objWorksheet.cell(row,col).alignment = \
        Alignment(horizontal='center', vertical='center')
        row+=1
    col+=1
objWorkbook.save("Student Table.xlsx")
